import openpyxl
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, BarChart3D, LineChart, Reference, Series

# Function to create a chart sheet for an original sheet
def create_chart_sheet(workbook, original_sheet):
    chart_sheet_name = f"{original_sheet}chart"
    chart_sheet = workbook.create_sheet(title=chart_sheet_name)
    return chart_sheet

# Load the Excel file containing the data
file_path = r"C:\Users\PMD - FEMI\Desktop\DELTA PROVINCE 20\ANNUAL\cleaned_files\cleaned_DELTA PROVINCE 20 ANNUAL.xlsx"
workbook = openpyxl.load_workbook(file_path)

# For MRR Charts
mrr_sheet_name = "MRR"
mrr_sheet = workbook[mrr_sheet_name]

# Define the data ranges for the MRR bar chart
first_timers_data = Reference(mrr_sheet, min_col=2, min_row=3, max_col=13, max_row=3)
attendance_data = Reference(mrr_sheet, min_col=2, min_row=4, max_col=13, max_row=4)

# Create a bar chart for MRR
mrr_chart = BarChart()
mrr_chart.type = "col"
mrr_chart.style = 10
mrr_chart.title = "2022/2023 MEMBERS RETENTION ANALYSIS"

# Add the data series to the MRR chart
first_timers_series = Series(first_timers_data, title="First Timers")
mrr_chart.append(first_timers_series)

attendance_series = Series(attendance_data, title="Avg. Attendance")
mrr_chart.append(attendance_series)

# Create a chart sheet for MRR and add the chart
mrr_chart_sheet = create_chart_sheet(workbook, mrr_sheet_name)
mrr_chart_location = mrr_chart_sheet.max_row + 2
mrr_chart_sheet.add_chart(mrr_chart, f'A{mrr_chart_location}')

# For CSR Charts
csr_sheet_name = "CSR DISTRIBUTION"  # Replace with the actual sheet name
csr_sheet = workbook[csr_sheet_name]

# Extract data for the bar chart (last values in each row - column N)
data = []
categories = []
for row in range(2, csr_sheet.max_row):
    categories.append(csr_sheet.cell(row=row, column=1).value)  # Assuming categories are in column A
    data.append(csr_sheet.cell(row=row, column=14).value)  # Column N contains the TOTAL values

# Create a 3D clustered column chart for CSR
csr_chart = BarChart3D()
csr_chart.style = 10
csr_chart.title = "2022/2023 CSR DISTRIBUTION"
csr_chart.x_axis.title = "CSR Categories"
csr_chart.y_axis.title = "Number of CSR Projects"
csr_chart.add_data(Reference(csr_sheet, min_col=14, min_row=2, max_row=csr_sheet.max_row - 1, max_col=14))
csr_chart.set_categories(Reference(csr_sheet, min_col=1, min_row=2, max_row=csr_sheet.max_row - 1, max_col=1))

# Create a chart sheet for CSR and add the chart
csr_chart_sheet = create_chart_sheet(workbook, csr_sheet_name)
csr_chart_location = csr_chart_sheet.max_row + 2
csr_chart_sheet.add_chart(csr_chart, f'A{csr_chart_location}')

# Add data labels to the CSR Distribution chart
data_labels = DataLabelList()
data_labels.showVal = True
csr_chart.dataLabels = data_labels

# For Church Analysis Chart
CA_sheet_name = "CHURCH ANALYSIS1"
CA_sheet = workbook[CA_sheet_name]

# Define the data range for all three columns (excluding the last row)
data_range_col1 = Reference(CA_sheet, min_col=1, min_row=1, max_row=CA_sheet.max_row-1)
data_range_col2 = Reference(CA_sheet, min_col=2, min_row=1, max_row=CA_sheet.max_row-1)
data_range_col3 = Reference(CA_sheet, min_col=3, min_row=1, max_row=CA_sheet.max_row-1)

# Create a combo chart with clustered columns and a line on the secondary axis
CA_chart = BarChart()
CA_chart.title = "CHURCH ANALYSIS"
CA_chart.style = 20
CA_chart.x_axis.title = "Attendance Range"
CA_chart.y_axis.title = "Total No. of Parishes"

# Add clustered columns to the primary axis
column_series = Series(data_range_col2, title_from_data=True)
column_series.graphicalProperties.solidFill = "0000FF"  # Set bar chart color to blue (Hex color code)
CA_chart.append(column_series)

# Create a line series on the secondary axis
line_chart = LineChart()
line_chart.y_axis.axId = 150
line_series = Series(data_range_col3, title_from_data=True)
line_series.graphicalProperties.solidFill = "FF0000"  # Set line chart color to red
line_chart.append(line_series)

# Combine the bar and line charts
CA_chart += line_chart

# Specify categories for the x-axis
categories = Reference(CA_sheet, min_col=1, min_row=2, max_row=CA_sheet.max_row-1)
CA_chart.set_categories(categories)

# Create a chart sheet for Church Analysis and add the chart
CA_chart_sheet = create_chart_sheet(workbook, CA_sheet_name)
CA_chart_location = CA_chart_sheet.max_row + 2
CA_chart_sheet.add_chart(CA_chart, f'A{CA_chart_location}')

# Save the modified workbook to the original file
workbook.save(file_path)

# Close the workbook
workbook.close()

# Print a statement once the charts are drawn
print("Charts added to the worksheet.")
