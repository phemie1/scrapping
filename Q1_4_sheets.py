import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import os, re

def clean_excel_file(file_path):
    """Cleans an Excel file at the given file path."""
    # Open the Excel file.
    wb = openpyxl.load_workbook(file_path)
    
    # Function to clean the "CHURCH ANALYSIS" sheet
    def clean_church_analysis_sheet(ws):
        # Load the sheet into a Pandas DataFrame, skipping the second row
        df = pd.read_excel(file_path, "CHURCH ANALYSIS", skiprows=[1])

        # Rename the columns
        new_columns = ["col_{}".format(i) for i in range(1, len(df.columns) + 1)]
        df.columns = new_columns

        # Drop specific columns
        columns_to_drop = ["col_1", "col_4", "col_5"]
        df = df.drop(columns=columns_to_drop)

        # Ensure there are at least 3 columns remaining
        while len(df.columns) < 3:
            df["dummy"] = None

        # Rename the columns explicitly to col1, col2, col3
        df.columns = ["col1", "col2", "col3"]

        # Rename specific columns
        df = df.rename(columns={"col1": "ATTENDANCE RANGE", "col2": "TOTAL NO. OF PARISHES", "col3": "% OF PARISHES"})

        # Check and delete rows where there is no data in TOTAL NO. OF PARISHES for the corresponding ATTENDANCE RANGE
        df = df.dropna(subset=["TOTAL NO. OF PARISHES"], how="all")

        # Move percentage values to the "% OF PARISHES" column
        df["% OF PARISHES"] = df["TOTAL NO. OF PARISHES"].str.extract(r'(\d+(?:\.\d+)?)%')

        # Remove brackets from values under TOTAL NO. OF PARISHES
        df["TOTAL NO. OF PARISHES"] = df["TOTAL NO. OF PARISHES"].str.extract(r'\((\d+)\)')

        # Convert the percentage strings to float
        df["% OF PARISHES"] = df["% OF PARISHES"].astype(float) / 100
        df["TOTAL NO. OF PARISHES"] = pd.to_numeric(df["TOTAL NO. OF PARISHES"])

        # Dynamically determine the last row by putting "TOTAL" in the "ATTENDANCE RANGE" column and summing the other two columns
        total_row = df.sum(numeric_only=True)
        # Round up to whole numbers
        total_row = total_row.round(0)
        total_row["ATTENDANCE RANGE"] = "TOTAL"
        df = pd.concat([df, total_row.to_frame().transpose()], ignore_index=True)
              
        # Save the modified DataFrame back to Excel with the desired format
        file_name = os.path.basename(file_path)
        sheet_name = "CHURCH ANALYSIS"
   
        # Church Analysis
        sheet_name = "CHURCH ANALYSIS"
        # Write the church analysis data to a new sheet
        church_analysis_sheet = wb.create_sheet(title=sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            church_analysis_sheet.append(row)

        # Get the max row number in the sheet
        max_row = church_analysis_sheet.max_row

        # Apply percentage format to the '% OF PARISHES' column
        for row_num in range(2, max_row + 1):
            church_analysis_sheet.cell(row=row_num, column=len(df.columns)).number_format = '0%'
    
        # Write the header row with formatting
        for col_num, value in enumerate(df.columns.values, start=1):
            cell = church_analysis_sheet.cell(row=1, column=col_num, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Apply "All Border" formatting to the entire worksheet
        for row in church_analysis_sheet.iter_rows(min_row=1, max_row=church_analysis_sheet.max_row, max_col=church_analysis_sheet.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                    right=openpyxl.styles.Side(style='thin'),
                                                    top=openpyxl.styles.Side(style='thin'),
                                                    bottom=openpyxl.styles.Side(style='thin'))

        # Set column width for 'A:B' and 'C:C'
        church_analysis_sheet.column_dimensions['A'].width = 15
        church_analysis_sheet.column_dimensions['B'].width = 15
        church_analysis_sheet.column_dimensions['C'].width = 15
    
        # Delete the original "CHURCH ANALYSIS" sheet from the cleaned workbook
        if "CHURCH ANALYSIS" in wb.sheetnames:
            del wb["CHURCH ANALYSIS"]
    
        # Print a statement once it finishes cleaning.
        print(f"Cleaning Excel File '{sheet_name}'...")

        # Save the cleaned DataFrame back to the sheet with the desired format
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Apply percentage format to the '% OF PARISHES' column
        for row_num in range(2, ws.max_row + 1):
            ws.cell(row=row_num, column=len(df.columns)).number_format = '0%'

    # Function to clean the "MRR" sheet
    def clean_mrr_sheet(ws):
        sheet_name = "MRR"  # Replace with the actual sheet name
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Set the height and width of row 1 and column A
            # ws.row_dimensions[1].height = 35.50
            ws.column_dimensions['A'].width = 33 

            # Rows to delete
            rows_to_delete = [1, 3, 9]

            # Delete specified rows in reverse order
            for row_number in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_number)
            
            # Modify row 5
            for cell in ws[5]:
                if cell.value:
                    cell.value = cell.value.split(" ", 1)[0]  # Keep only the first value before the space

            # Clean all cells in row 6 with the specified format
            for cell in ws[6]:
                if cell.value:
                    matches = re.findall(r'(-?\d+\.\d+%)(?:\s*\([\d\./]*\)\s*\*\s*100)?', cell.value)
                    if matches:
                        cleaned_value = " ".join(matches)
                        cell.value = cleaned_value

            # Clean "peopleinitial" in row 4
            for row in ws.iter_rows(min_row=2, max_row=4):
                for cell in row:
                    if cell.value:
                        cell.value = cell.value.replace("Peopleinitial","")

            # Clean "people" in row 2, 3, and 4
            for row in ws.iter_rows(min_row=2, max_row=4):
                for cell in row:
                    if cell.value:
                        cell.value = cell.value.replace("People","")
            
            # Iterate through all rows and columns in the specified range from B2 to M2
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4, values_only=True), start=2):
                for col_index, cell_value in enumerate(row, start=2):
                    # Check if the cell content is text or numeric
                    if isinstance(cell_value, str):
                        try:
                            numeric_value = float(cell_value.replace(',', ''))  # Convert text to numeric
                            ws.cell(row=row_index, column=col_index, value=numeric_value)
                        except ValueError:
                            pass  # Keep the original value if conversion fails
                    elif isinstance(cell_value, (int, float)):
                        pass  # The cell already contains a numeric value
                    else:
                        pass  # The cell is empty
            
            # Apply "All Border" formatting to the entire worksheet.
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                        right=openpyxl.styles.Side(style='thin'),
                                                        top=openpyxl.styles.Side(style='thin'),
                                                        bottom=openpyxl.styles.Side(style='thin'))

             # Set the specified text in cells A1 to A6
            text_in_column_A = ["DETAILS", "TOTAL NO OF FIRST TIMERS(MONTHLY)", "FIRST TIMERS (CUMMULATIVE)",
                                "AVG. ATTENDANCE", "ATTENDANCE CHANGE", "MEMBERS RETENTION RATE"]
            for i, text in enumerate(text_in_column_A, start=1):
                ws[f'A{i}'] = text
    
    # Function to clean the "CSR" sheet
    def clean_csr_sheet(ws):
        sheet_name = "CSR" 
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Rows to delete
            rows_to_delete = [1, 3, 6, 7, 12]
            # Delete specified rows in reverse order
            for row_number in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_number)

            # Set the height and width of row 1 and column A
            # ws.row_dimensions[1].height = 35.50
            ws.column_dimensions['A'].width = 37
            ws.column_dimensions['N'].width = 11 

            # Set the specified text in cells A1 to A7
            text_in_column_A = ["DETAILS", "TOTAL NUMBER OF CSR PROJECTS", "TOTAL ESTIMATED FINANCIAL EXPENDITURE", 
                                "COMPLIANCE", "NUMBER OF BENEFICIARIES", "NUMBER OF BENEFICIARY LGAs", "BENEFICIARY STATE(S)"]
            for i, text in enumerate(text_in_column_A, start=1):
                ws[f'A{i}'] = text
            
            # Iterate through all rows and columns in the specified range from B2 to M2
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5, values_only=True), start=2):
                for col_index, cell_value in enumerate(row, start=2):
                    # Check if the cell content is text or numeric
                    if isinstance(cell_value, str):
                        try:
                            numeric_value = float(cell_value.replace(',', ''))  # Convert text to numeric
                            ws.cell(row=row_index, column=col_index, value=numeric_value)
                        except ValueError:
                            pass  # Keep the original value if conversion fails
                    elif isinstance(cell_value, (int, float)):
                        pass  # The cell already contains a numeric value
                    else:
                        pass  # The cell is empty
            
            # Apply "All Border" formatting to the entire worksheet.
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                        right=openpyxl.styles.Side(style='thin'),
                                                        top=openpyxl.styles.Side(style='thin'),
                                                        bottom=openpyxl.styles.Side(style='thin'))

    # Function to clean the "CSR DISTRIBUTION" sheet
    def clean_csr_distribution_sheet(ws):
        sheet_name = "CSR DISTRIBUTION"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Rows to delete
            rows_to_delete = [1, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32,
                            33, 34, 35, 36, 37, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 54, 55, 56, 57, 58, 60, 61, 62,
                            63, 64, 66, 67, 68, 69, 70, 71]

            # Delete specified rows in reverse order
            for row_number in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_number)

            # Apply bold and wrap text to the header row.
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=1, column=col)
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(wrap_text=True)

            # Set the height and width of row 1 and column A
            ws.row_dimensions[1].height = 35.50
            ws.column_dimensions['A'].width = 32
            
            # Iterate through all rows and columns in the specified range from B2 to M2
            for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5, values_only=True), start=2):
                for col_index, cell_value in enumerate(row, start=2):
                    # Check if the cell content is text or numeric
                    if isinstance(cell_value, str):
                        try:
                            numeric_value = float(cell_value.replace(',', ''))  # Convert text to numeric
                            ws.cell(row=row_index, column=col_index, value=numeric_value)
                        except ValueError:
                            pass  # Keep the original value if conversion fails
                    elif isinstance(cell_value, (int, float)):
                        pass  # The cell already contains a numeric value
                    else:
                        pass  # The cell is empty
            
            # Create a pandas DataFrame from the worksheet
            data = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                data.append(row)
            # Convert the data to a pandas DataFrame
            df = pd.DataFrame(data)
            # Exclude the first column (index 0) before applying pd.to_numeric
            df.iloc[:, 1:] = df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
            # Calculate the column sums excluding the first column
            column_sums = df.iloc[:, 1:].sum()
            # Insert the column sums in a new row at the end
            ws.append(['TOTAL'] + list(column_sums))

            # Apply "All Border" formatting to the entire worksheet.
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                        right=openpyxl.styles.Side(style='thin'),
                                                        top=openpyxl.styles.Side(style='thin'),
                                                        bottom=openpyxl.styles.Side(style='thin'))

            # Align columns to center
            for col in range(2, ws.max_column + 1):
                col_letter = get_column_letter(col)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Apply bold to column N
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
                for cell in row:
                    cell.font = Font(bold=True)

            # Apply bold to row 10
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=10, max_row=10):
                for cell in col:
                    cell.font = Font(bold=True)

    # List of sheet names to clean
    sheets_to_clean = ["CHURCH ANALYSIS", "MRR", "CSR", "CSR DISTRIBUTION"]

    for sheet_name in sheets_to_clean:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Call the appropriate cleaning function based on the sheet name
            if sheet_name == "CHURCH ANALYSIS":
                clean_church_analysis_sheet(ws)
            elif sheet_name == "MRR":
                clean_mrr_sheet(ws)
            elif sheet_name == "CSR":
                clean_csr_sheet(ws)
            elif sheet_name == "CSR DISTRIBUTION":
                clean_csr_distribution_sheet(ws)

    # Save the modified workbook (original file) after cleaning all specified sheets
    wb.save(file_path)

    # Print a statement once it finishes cleaning.
    print(f"Cleaning Excel File '{file_path}'...")

# Specify the folder path containing Excel files
folder_path =r"C:\Users\PMD - FEMI\Desktop\DELTA PROVINCE 20\QUARTER 1\cleaned_files"

# Iterate over all Excel files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        clean_excel_file(file_path)
