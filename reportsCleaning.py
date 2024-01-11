import openpyxl
import os,re
import glob
import xlsxwriter as xl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from collections import OrderedDict 
from openpyxl.utils.dataframe import dataframe_to_rows


def clean_excel_file(file_path, output_directory, cleaned_file_path):
    """Cleans an Excel file at the given file path."""
    global file_name
    # Open the Excel file.
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in ["AVG. ATTENDANCE", "HOUSE FELLOWSHIP", "CONVERTS"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Setting row's height and columns's widths
            ws.row_dimensions[1].height = 55
            ws.column_dimensions['A'].width = 3.50
            ws.column_dimensions['B'].width = 33
            ws.column_dimensions['C'].width = 15.50 
            ws.column_dimensions['P'].width = 13
            ws.column_dimensions['Q'].width = 12

            # Delete rows 1 to 3, and make row 4 the new header.
            ws.delete_rows(1, 3)

            # Modify the header row in columns D to O.
            for col in range(4, 16):
                header_cell = ws.cell(row=1, column=col)
                header_value = header_cell.value
                if header_value:
                    # Extract the month and year by splitting the header value at space and keeping the last two parts.
                    month_year = ' '.join(header_value.split(' ')[-2:])
                    header_cell.value = month_year

            # Create a list of all rows except for row 1 (new header).
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            # Find and delete any row with any blank cell from A to P.
            rows_to_delete = [i for i, row in enumerate(rows) if any(cell is None for cell in row[:16])]

            for i in reversed(rows_to_delete):
                ws.delete_rows(i + 2, 1)

            # Delete columns Q to R.
            ws.delete_cols(17, 18 + 1)

            # Set headers in specific cells.
            ws['A1'] = 'S/N'
            ws['B1'] = 'NAME OF PARISH'

            # Replace all instances of '-' with '0' in all cells, except for the header row.
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row, col).value:
                        # Check if the cell value is a string.
                        if isinstance(ws.cell(row, col).value, str):
                            # Check if the cell value is a valid integer.
                            if ws.cell(row, col).value.isnumeric():
                                ws.cell(row, col).value = int(ws.cell(row, col).value)
                            # Replace all instances of '-' with '0' in the cell value.
                            ws.cell(row, col).value = str(ws.cell(row, col).value).replace("-", "0")
           
            if sheet_name == "AVG. ATTENDANCE":
              # Calculate and set the '% GROWTH' column values.
              for row in range(2, ws.max_row + 1):
                  p_cell = f'P{row}'
                  c_cell = f'C{row}'
                  q_cell = f'Q{row}'
                  # Use the IFERROR() function to handle the case where the denominator is 0.
                  ws[q_cell] = f'=ROUND(IFERROR(({p_cell} - {c_cell}) / {c_cell}, 0), 4)'
                  ws['Q1'] = 'GROWTH PERCENTAGE'
          
            elif sheet_name in ["HOUSE FELLOWSHIP", "CONVERTS"]:
              # Populate column Q with row numbers.
              for row in range(2, ws.max_row + 1):
                  q_cell = f'Q{row}'
                  ws[q_cell] = f'{row - 1}{"st" if row == 2 else "nd" if row == 3 else "rd" if row == 4 else "th"}'
                  ws['Q1'] = 'RANKING'

              #sort_sheet_by_column(ws, 16)  # Sort by column P (index 15)
            # Calculate the last row with data in column C.
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 3).value is None:
                    last_row = row - 1
                    break
            else:
                last_row = ws.max_row  # If no empty cell is found, use the last row. 

            # Convert cell values to numbers before calculating the sum.
            for row in range(2, last_row + 1):
                for col in range(3, 17):
                    cell_value = ws.cell(row, col).value
                    if isinstance(cell_value, str) and cell_value.isnumeric():
                        ws.cell(row, col).value = int(cell_value)
            
            # Calculate the sum of each column from C2 to the last row with data and place the results in a straight row.
            result_row = last_row + 1  # Determine the row to place the results.
            # Sum up columns C to P (3 to 17)
            for col in range(3, 17):  
              col_letter = get_column_letter(col)
              sum_formula = f'=SUM({col_letter}2:{col_letter}{last_row})'
              ws[f'{col_letter}{result_row}'] = sum_formula
              ws[f'{col_letter}{result_row}'].font = Font(bold=True)
              ws[f'{col_letter}{result_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

           
            # Apply bold and wrap text to the header row.
            for col in range(1, ws.max_column + 1):
              header_cell = ws.cell(row=1, column=col)
              header_cell.font = Font(bold=True)
              header_cell.alignment = Alignment(wrap_text=True)

            # Iterate through all columns (except A and B) and align them to the center.
            for col in range(3, ws.max_column + 1):
                col_letter = get_column_letter(col)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply "All Border" formatting to the entire worksheet.
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                       right=openpyxl.styles.Side(style='thin'),
                                                       top=openpyxl.styles.Side(style='thin'),
                                                       bottom=openpyxl.styles.Side(style='thin'))   
    
    # Load the Excel file with the specific sheet name, skipping the second row
    df = pd.read_excel(file_path, "CHURCH ANALYSIS", skiprows=[1])

    # Rename the columns
    new_columns = ["col_{}".format(i) for i in range(1, len(df.columns) + 1)]
    df.columns = new_columns

    # Drop specific columns
    columns_to_drop = ["col_1", "col_4", "col_5"]
    df = df.drop(columns=columns_to_drop)

    # Ensure there are at least 3 columns remaining
    while len(df.columns) < 3:
        df["dummy"] = None

    # Rename the columns explicitly to col1, col2, col3
    df.columns = ["col1", "col2", "col3"]

    # Rename specific columns
    df = df.rename(columns={"col1": "ATTENDANCE RANGE", "col2": "TOTAL NO. OF PARISHES", "col3": "% OF PARISHES"})

    # Check and delete rows where there is no data in TOTAL NO. OF PARISHES for the corresponding ATTENDANCE RANGE
    df = df.dropna(subset=["TOTAL NO. OF PARISHES"], how="all")

    # Move percentage values to the "% OF PARISHES" column
    df["% OF PARISHES"] = df["TOTAL NO. OF PARISHES"].str.extract(r'(\d+\.\d+)%')
    # Remove brackets from values under TOTAL NO. OF PARISHES
    df["TOTAL NO. OF PARISHES"] = df["TOTAL NO. OF PARISHES"].str.extract(r'\((\d+)\)')

    # Convert the percentage strings to float
    df["% OF PARISHES"] = df["% OF PARISHES"].astype(float) / 100
    df["TOTAL NO. OF PARISHES"] = pd.to_numeric(df["TOTAL NO. OF PARISHES"])

    # Dynamically determine the last row by putting "TOTAL" in the "ATTENDANCE RANGE" column and summing the other two columns
    total_row = df.sum(numeric_only=True)
    # Round up to whole numbers
    total_row = total_row.round(0)
    total_row["ATTENDANCE RANGE"] = "TOTAL"
    df = pd.concat([df, total_row.to_frame().transpose()], ignore_index=True)
              
    # Save the modified DataFrame back to Excel with the desired format
    file_name = os.path.basename(file_path)
    sheet_name = "CHURCH ANALYSIS"
   
    # Church Analysis
    sheet_name = "CHURCH ANALYSIS"
    # Write the church analysis data to a new sheet
    church_analysis_sheet = wb.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        church_analysis_sheet.append(row)

    # Get the max row number in the sheet
    max_row = church_analysis_sheet.max_row

    # Apply percentage format to the '% OF PARISHES' column
    for row_num in range(2, max_row + 1):
        church_analysis_sheet.cell(row=row_num, column=len(df.columns)).number_format = '0.00%'
    
    # Write the header row with formatting
    for col_num, value in enumerate(df.columns.values, start=1):
        cell = church_analysis_sheet.cell(row=1, column=col_num, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Apply "All Border" formatting to the entire worksheet
    for row in church_analysis_sheet.iter_rows(min_row=1, max_row=church_analysis_sheet.max_row, max_col=church_analysis_sheet.max_column):
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                right=openpyxl.styles.Side(style='thin'),
                                                top=openpyxl.styles.Side(style='thin'),
                                                bottom=openpyxl.styles.Side(style='thin'))

    # Set column width for 'A:B' and 'C:C'
    church_analysis_sheet.column_dimensions['A'].width = 15
    church_analysis_sheet.column_dimensions['B'].width = 15
    church_analysis_sheet.column_dimensions['C'].width = 15
    
    # Delete the original "CHURCH ANALYSIS" sheet from the cleaned workbook
    if "CHURCH ANALYSIS" in wb.sheetnames:
        del wb["CHURCH ANALYSIS"]
    
    # Print a statement once it finishes cleaning.
    print(f"Cleaning Excel File '{sheet_name}'...")
    
    
    # 1: Specify the sheet name
    sheet_name = "MRR"  # Replace with the actual sheet name

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Set the height and width of row 1 and column A
        #ws.row_dimensions[1].height = 35.50
        ws.column_dimensions['A'].width = 33 

        # Rows to delete
        rows_to_delete = [1, 3, 9]

        # Delete specified rows in reverse order
        for row_number in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_number)

        # Set the specified text in cells A1 to A6
        text_in_column_A = ["DETAILS", "TOTAL NO OF FIRST TIMERS(MONTHLY)", "FIRST TIMERS (CUMMULATIVE)",
                            "AVG. ATTENDANCE", "ATTENDANCE CHANGE", "MEMBERS RETENTION RATE"]
        for i, text in enumerate(text_in_column_A, start=1):
            ws[f'A{i}'] = text

        # Modify row 5
        for cell in ws[5]:
            if cell.value:
                cell.value = cell.value.split(" ", 1)[0]  # Keep only the first value before the space

        # Clean all cells in row 6 with the specified format
        for cell in ws[6]:
            if cell.value:
                matches = re.findall(r'(-?\d+\.\d+%)(?:\s*\([\d\./]*\)\s*\*\s*100)?', cell.value)
                if matches:
                    cleaned_value = " ".join(matches)
                    cell.value = cleaned_value

        # Clean "peopleinitial" in row 4
        for row in ws.iter_rows(min_row=2, max_row=4):
            for cell in row:
                if cell.value:
                    cell.value = cell.value.replace("Peopleinitial","")

        # Clean "people" in row 2, 3, and 4
        for row in ws.iter_rows(min_row=2, max_row=4):
            for cell in row:
                if cell.value:
                    cell.value = cell.value.replace("People","")
        
        # Iterate through all rows and columns in the specified range from B2 to M2
        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=13, values_only=True), start=2):
            for col_index, cell_value in enumerate(row, start=2):
                # Check if the cell content is text or numeric
                if isinstance(cell_value, str):
                    try:
                        numeric_value = float(cell_value.replace(',', ''))  # Convert text to numeric
                        ws.cell(row=row_index, column=col_index, value=numeric_value)
                    except ValueError:
                        pass  # Keep the original value if conversion fails
                elif isinstance(cell_value, (int, float)):
                    pass  # The cell already contains a numeric value
                else:
                    pass  # The cell is empty
        
        # Apply "All Border" formatting to the entire worksheet.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                    right=openpyxl.styles.Side(style='thin'),
                                                    top=openpyxl.styles.Side(style='thin'),
                                                    bottom=openpyxl.styles.Side(style='thin'))

    # 2: Specify the sheet name
    sheet_name = "CSR" 

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Rows to delete
        rows_to_delete = [1, 3, 6, 7, 12]

        # Delete specified rows in reverse order
        for row_number in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_number)

         # Set the height and width of row 1 and column A
        #ws.row_dimensions[1].height = 35.50
        ws.column_dimensions['A'].width = 37
        ws.column_dimensions['N'].width = 11 

        # Set the specified text in cells A1 to A7
        text_in_column_A = ["DETAILS", "TOTAL NUMBER OF CSR PROJECTS", "TOTAL ESTIMATED FINANCIAL EXPENDITURE", 
                            "COMPLIANCE", "NUMBER OF BENEFICIARIES", "NUMBER OF BENEFICIARY LGAs", "BENEFICIARY STATE(S)"]
        for i, text in enumerate(text_in_column_A, start=1):
            ws[f'A{i}'] = text
        
        # Iterate through all rows and columns in the specified range from B2 to M2
        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=14, values_only=True), start=2):
            for col_index, cell_value in enumerate(row, start=2):
                # Check if the cell content is text or numeric
                if isinstance(cell_value, str):
                    try:
                        numeric_value = float(cell_value.replace(',', ''))  # Convert text to numeric
                        ws.cell(row=row_index, column=col_index, value=numeric_value)
                    except ValueError:
                        pass  # Keep the original value if conversion fails
                elif isinstance(cell_value, (int, float)):
                    pass  # The cell already contains a numeric value
                else:
                    pass  # The cell is empty
        
        # Apply "All Border" formatting to the entire worksheet.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                    right=openpyxl.styles.Side(style='thin'),
                                                    top=openpyxl.styles.Side(style='thin'),
                                                    bottom=openpyxl.styles.Side(style='thin'))

    # 3: Specify the sheet name
    sheet_name = "CSR DISTRIBUTION"

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Rows to delete
        rows_to_delete = [1, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32,
                        33, 34, 35, 36, 37, 38, 40, 41, 42, 43, 44, 45, 46, 48, 49, 50, 51, 52, 54, 55, 56, 57, 58, 60, 61, 62,
                        63, 64, 66, 67, 68, 69, 70, 71]

        # Delete specified rows in reverse order
        for row_number in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_number)

        # Apply bold and wrap text to the header row.
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(wrap_text=True)

        # Set the height and width of row 1 and column A
        ws.row_dimensions[1].height = 35.50
        ws.column_dimensions['A'].width = 32
        
         # Iterate through all rows and columns in the specified range from B2 to M2
        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=14, values_only=True), start=2):
            for col_index, cell_value in enumerate(row, start=2):
                # Check if the cell content is text or numeric
                if isinstance(cell_value, str):
                    try:
                        numeric_value = float(cell_value.replace(',', ''))  # Convert text to numeric
                        ws.cell(row=row_index, column=col_index, value=numeric_value)
                    except ValueError:
                        pass  # Keep the original value if conversion fails
                elif isinstance(cell_value, (int, float)):
                    pass  # The cell already contains a numeric value
                else:
                    pass  # The cell is empty
        
        # Create a pandas DataFrame from the worksheet
        data = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            data.append(row)
        # Convert the data to a pandas DataFrame
        df = pd.DataFrame(data)
        df = df.apply(pd.to_numeric, errors='coerce')
        # Calculate the column sums
        column_sums = df.sum()
        # Insert the column sums in a new row at the end
        ws.append(list(column_sums))

        # Apply "All Border" formatting to the entire worksheet.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                    right=openpyxl.styles.Side(style='thin'),
                                                    top=openpyxl.styles.Side(style='thin'),
                                                    bottom=openpyxl.styles.Side(style='thin'))

        # Align columns to center
        for col in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col)
            for row in range(1, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Apply bold to column N
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=14, max_col=14):
            for cell in row:
                cell.font = Font(bold=True)

        # Apply bold to row 10
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=10, max_row=10):
            for cell in col:
                cell.font = Font(bold=True)
   
   # Define the file name for the cleaned Excel file.
    file_name = os.path.basename(file_path)
    cleaned_file_name = f'cleaned_{file_name}'

    # Construct the full path for the cleaned file in the new folder
    cleaned_file_path = os.path.join(output_directory, cleaned_file_name)

    # Save the cleaned workbook to the new folder
    wb.save(cleaned_file_path)

    # Print a statement once it finishes cleaning.
    print(f"Cleaning Excel File '{file_name}'...")

def clean_excel_files_in_folder(folder_path, output_directory):
    """Cleans all Excel files in the given folder and saves the cleaned files in the specified output directory."""
    # Create the "cleaned" folder if it doesn't exist
    cleaned_folder_path = os.path.join(output_directory, f"cleaned_{os.path.basename(folder_path)}")
    os.makedirs(cleaned_folder_path, exist_ok=True)

    for file_path in glob.glob(folder_path + "/*.xlsx"):
        # Extract the file name and extension
        file_name, file_extension = os.path.splitext(os.path.basename(file_path))
        
        # Construct the cleaned file name
        cleaned_file_name = f'cleaned_{file_name}{file_extension}'
        
        # Construct the full path for the cleaned file in the new folder
        cleaned_file_path = os.path.join(cleaned_folder_path, cleaned_file_name)

        # Call the clean_excel_file function
        clean_excel_file(file_path, cleaned_folder_path, cleaned_file_path)

   
# Specify the folder that contains the Excel files to clean
input_folder_path = r"C:\Users\PMD - FEMI\Desktop\PROVINCIAL REPORTS\REGIONS"

# Specify the output directory for the cleaned files
output_directory = r"C:\Users\PMD - FEMI\Desktop\PROVINCIAL REPORTS\REGIONS"

# Clean Excel files in the input folder and save them to the output directory
clean_excel_files_in_folder(input_folder_path, output_directory)
