import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def process_excel_files(folder_path, output_excel_path):
    # Create empty DataFrames to store the combined data
    combined_data_avg_quarter = pd.DataFrame()
    combined_data_benchmark = pd.DataFrame()

    # Loop through all files in the specified folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)

            # Read the Excel file and extract the Sheet1
            try:
                df = pd.read_excel(file_path, sheet_name="Sheet1")
            except Exception as e:
                print(f"Error reading {file_path}: {e}")
                continue

            # Check if the required columns are present in the sheet
            if 'BENCHMARK (Avg. OF SEP. 22 - AUG. 23)' in df.columns and 'Avg. QUARTER 1' in df.columns:
                # Filter rows where 'Avg. QUARTER 1' values are less than 100
                filtered_df_avg_quarter = df[df['Avg. QUARTER 1'] < 100].copy()
                filtered_df_avg_quarter['Province'] = os.path.splitext(filename)[0]  # Use file name as province name
                combined_data_avg_quarter = pd.concat([combined_data_avg_quarter, filtered_df_avg_quarter[['Province', 'Avg. QUARTER 1']]], ignore_index=True)

                # Filter rows where 'BENCHMARK' values are less than 100
                filtered_df_benchmark = df[df['BENCHMARK (Avg. OF SEP. 22 - AUG. 23)'] < 100].copy()
                filtered_df_benchmark['Province'] = os.path.splitext(filename)[0]  # Use file name as province name
                combined_data_benchmark = pd.concat([combined_data_benchmark, filtered_df_benchmark[['Province', 'BENCHMARK (Avg. OF SEP. 22 - AUG. 23)']]], ignore_index=True)

    # Check if there is data with values less than 100 for 'Avg. QUARTER 1'
    if not combined_data_avg_quarter.empty:
        # Count occurrences of each province for 'Avg. QUARTER 1' values less than 100
        count_df_avg_quarter = combined_data_avg_quarter.groupby('Province').size().reset_index(name='NOV-23')
        # Save count DataFrame for 'Avg. QUARTER 1' to a temporary Excel file
        count_df_avg_quarter.to_excel(os.path.join(output_excel_path, 'count_avg_quarter_temp.xlsx'), index=False)
        #print("Count for 'Avg. QUARTER 1' saved to count_avg_quarter_temp.xlsx")
    else:
        print("No data with 'Avg. QUARTER 1' values less than 100.")

    # Check if there is data with values less than 100 for 'BENCHMARK'
    if not combined_data_benchmark.empty:
        # Count occurrences of each province for 'BENCHMARK' values less than 100
        count_df_benchmark = combined_data_benchmark.groupby('Province').size().reset_index(name='SEP-22/AUG-23')
        # Save count DataFrame for 'BENCHMARK' to a temporary Excel file
        count_df_benchmark.to_excel(os.path.join(output_excel_path, 'count_benchmark_temp.xlsx'), index=False)
        #print("Count for 'BENCHMARK (Avg. OF SEP. 22 - AUG. 23)' saved to count_benchmark_temp.xlsx")
    else:
        print("No data with 'BENCHMARK' values less than 100.")

    # Read both temporary Excel files
    count_df_avg_quarter = pd.read_excel(os.path.join(output_excel_path, 'count_avg_quarter_temp.xlsx'))
    count_df_benchmark = pd.read_excel(os.path.join(output_excel_path, 'count_benchmark_temp.xlsx'))

    # Merge the two DataFrames on 'Province'
    merged_df = pd.merge(count_df_benchmark, count_df_avg_quarter, on='Province', how='outer')

    # Save the merged DataFrame to the final Excel file with additional columns
    folder_name = os.path.basename(folder_path)
    result_df = pd.merge(count_df_benchmark, count_df_avg_quarter, on='Province', how='outer').fillna(0)

    # Add a 'TOTAL' row with the sum of each column
    total_row = result_df.sum(numeric_only=True)
    total_row['Province'] = 'TOTAL'

    # Use concat instead of append for adding the 'TOTAL' row
    result_df_final = pd.concat([result_df, total_row.to_frame().transpose()], ignore_index=True)

    # Calculate the difference and add the 'DIFF' column
    result_df_final['DIFF'] = result_df_final['SEP-22/AUG-23'] - result_df_final['NOV-23']

    # Reorder columns as per the requirement
    result_df_final = result_df_final[['Province', 'SEP-22/AUG-23', 'NOV-23', 'DIFF']]

    # Save to Excel
    excel_file_path = os.path.join(output_excel_path, f'{folder_name}_pivot.xlsx')
    result_df_final.to_excel(excel_file_path, index=False)

    # Bold the 'TOTAL' row
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # Find the index of the 'TOTAL' row
    total_row_index = result_df_final.index[result_df_final['Province'] == 'TOTAL'][0] + 2  # Add 2 for 1-indexed Excel rows

    # Apply bold formatting to the 'TOTAL' row
    for row in sheet.iter_rows(min_row=total_row_index, max_row=total_row_index, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = Font(bold=True)


    # Save the changes to the Excel file
    wb.save(excel_file_path)

    print(f"Merged DataFrames saved to {folder_name}_pivot.xlsx")

    # Remove temporary files
    temp_files = ['count_avg_quarter_temp.xlsx', 'count_benchmark_temp.xlsx']
    for temp_file in temp_files:
        try:
            os.remove(os.path.join(output_excel_path, temp_file))
        except Exception as e:
            print(f"Error removing temporary file {temp_file}: {e}")

if __name__ == "__main__":
    folder_path = r'C:\Users\PMD - FEMI\Desktop\pivot\Redemption City Region\Redemption City Region'  # Change this to the actual folder path
    output_excel_path = r'C:\Users\PMD - FEMI\Desktop\pivot\MergedRegions_pivot'  # Change this to the desired output Excel file path
    process_excel_files(folder_path, output_excel_path)
