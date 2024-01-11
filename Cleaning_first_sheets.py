import pandas as pd
import os
import openpyxl
from openpyxl.styles import Border, Side

# Specify the directory where your Excel files are located
excel_directory = r"C:\Users\PMD - FEMI\Desktop\Tiers\Tier 1"
# Get a list of all Excel files in the specified directory
excel_files = [file for file in os.listdir(excel_directory) if file.endswith('.xlsx')]

# Function to clean the first three sheets
def clean_first_three_sheets(df):
    # Remove the first and second rows]
    df_cleaned = df.drop([0, 1])

    # Remove the last four columns
    df_cleaned = df_cleaned.iloc[:, :-4]

    # Define your custom header
    custom_header = ['S/N', 'PARISH NAME', 'BENCHMARK (AUG. 2023.)', 'SEP-23', 'OCT-23', 'NOV-23', 'AVG.SEP-23 - NOV-23']
    # Set the custom header to the DataFrame
    df_cleaned.columns = custom_header
    # Remove rows where any of the columns (including the first two) are empty
    df_cleaned = df_cleaned.dropna(subset=custom_header, how='any')
   
    # Convert specific columns from text to numerical values
    numeric_columns = df_cleaned.columns[2:]  # Assuming the first two columns are not numeric
    df_cleaned[numeric_columns] = df_cleaned[numeric_columns].apply(pd.to_numeric, errors='coerce')
    # Fill missing values (NaN) with zero in numeric columns
    df_cleaned[numeric_columns] = df_cleaned[numeric_columns].fillna(0)
    # Sort the DataFrame by 'Avg Sep. 22 - Aug. 23' in descending order
    df_sorted = df_cleaned.sort_values(by=['AVG.SEP-23 - NOV-23'], ascending=False, ignore_index=True)

    # Add the sum row under the 'Parish' column
    sum_row = pd.DataFrame(df_sorted.iloc[:, 2:].sum(axis=0)).T
    sum_row['S/N'] = ''
    sum_row['PARISH NAME'] = 'TOTAL'
    sum_row = sum_row[df_sorted.columns]  # Reorder columns to match the original DataFrame

    # Concatenate the sum row to the sorted DataFrame
    df_result = pd.concat([df_sorted, sum_row], ignore_index=True)
    return df_result

# Function to add the '% GROWTH' column
def add_percent_growth_column(df):
    # Calculate percentage growth and add a new column
    df['% GROWTH'] = ((df['AVG.SEP-23 - NOV-23'] - df['BENCHMARK (AUG. 2023.)']) / df['AVG.SEP-23 - NOV-23']) * 100

    # Format the '% Growth' column as a percentage
    df['% GROWTH'] = df['% GROWTH'].apply(lambda x: f'{x:.2f}%')
    return df

# Function to get the rank suffix
def get_rank_suffix(rank):
    if 10 <= rank % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(rank % 10, 'th')
    return f'{rank}{suffix}'

# Function to add the 'RANK' column
def add_ranking_column(df):
    # Add the 'RANK' column based on '% GROWTH' column
    df['RANKING'] = df['% GROWTH'].rank(ascending=False, method='min').astype(int)

    # Add the rank suffix to the 'RANK' column
    df['RANKING'] = df['RANKING'].apply(get_rank_suffix)

    return df

# Loop through each Excel file and clean the sheets
for file in excel_files:
    # Read the Excel file into a dictionary of DataFrames (one DataFrame for each sheet)
    excel_path = os.path.join(excel_directory, file)
    all_sheets_dict = pd.read_excel(excel_path, sheet_name=None)

    # Clean the first three sheets in the current Excel file
    for sheet_name in list(all_sheets_dict.keys())[:3]:
        all_sheets_dict[sheet_name] = clean_first_three_sheets(all_sheets_dict[sheet_name])

    # Create a new directory for the cleaned files
    cleaned_directory = os.path.join(excel_directory, "cleaned_files")
    os.makedirs(cleaned_directory, exist_ok=True)

    # Create a new filename for the cleaned file
    cleaned_file_name = f"cleaned_{file}"
    cleaned_file_path = os.path.join(cleaned_directory, cleaned_file_name)

  # Save the cleaned sheets to the new Excel file
with pd.ExcelWriter(cleaned_file_path, engine='openpyxl') as writer:
    for sheet_name, df in all_sheets_dict.items():
        # Check if the DataFrame is not empty before writing to Excel
        if not df.empty:
            if sheet_name == 'AVG. ATTENDANCE':
                df = add_percent_growth_column(df)
            elif sheet_name in ['CONVERTS', 'HOUSE FELLOWSHIP']:
                df = add_ranking_column(add_percent_growth_column(df))  # Add % GROWTH before ranking

            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Access the worksheet to apply cell borders
            ws = writer.sheets[sheet_name]
            # Apply cell borders
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin'))

print(f"Cleaning completed.")
